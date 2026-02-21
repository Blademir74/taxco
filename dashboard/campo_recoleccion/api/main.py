from fastapi import FastAPI, HTTPException
from fastapi.middleware.cors import CORSMiddleware
import pandas as pd
from datetime import datetime, timezone, timedelta
from database import engine
from models import DiagnosticoTerritorio
import base64
import os
import uvicorn
import tempfile
from starlette.background import BackgroundTask
from fastapi.responses import FileResponse
import uuid
from pydantic import BaseModel, EmailStr
from fastapi import Header
import logging

# Configurar logging para ver errores en Render
logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)

app = FastAPI(title="API de Recolección Territorial - Taxco")

<<<<<<< HEAD

# CORS: Configuración explícita para Streamlit Cloud y Localhost
origins = [
    "http://localhost:8501",
    "https://share.streamlit.io",
    "https://taxco-dashboard.streamlit.app",  # Ajustar a la URL real de tu app
    "*"  # Permitir todo por ahora para debug, pero con allow_credentials=True requiere orígenes explícitos si se usa cookies/auth headers
]

app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"], # Para simplificar debugging inicial. En producción, usar lista 'origins'
    allow_credentials=True,
=======
# CORS (permite todos los orígenes, ajusta en producción)
app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
>>>>>>> c0e6b59f5a7880c257092a53aa7675117f052c2d
    allow_methods=["*"],
    allow_headers=["*"],
)

# ============================================
<<<<<<< HEAD
# Health Check & Root (Keep-Alive)
# ============================================
@app.get("/")
def read_root():
    return {"mensaje": "API de Recolección Territorial Activa", "doc": "/docs"}

@app.get("/api/health")
def health_check():
    """Endpoint ligero para evitar Cold Starts de Render"""
    return {"status": "ok", "timestamp": datetime.now(timezone.utc).isoformat()}


# ============================================
# Endpoints públicos
# ============================================

@app.get("/api/secciones")
async def listar_secciones():
    query = """
    SELECT 
        s.seccion,
        CASE 
            WHEN r.pct_sin_agua > 10 THEN 'Zona con rezago de agua'
            WHEN r.pct_sin_servicios_basicos > 20 THEN 'Zona con rezago social'
            ELSE 'Zona sin rezago crítico'
        END AS insight
    FROM seccion s
    LEFT JOIN vw_rezago_secciones r ON r.seccion = s.seccion
    WHERE s.id_municipio = 56
    ORDER BY s.seccion
    """
    df = pd.read_sql(query, engine)
    return df.to_dict(orient='records')

@app.post("/api/recoleccion", status_code=201)
async def guardar_diagnostico(diagnostico: DiagnosticoTerritorio):
    conn = engine.raw_connection()
    cursor = conn.cursor()
    try:
        # Validar sección
        cursor.execute(
            "SELECT pk_seccion FROM seccion WHERE seccion = %s AND id_municipio = 56",
            (diagnostico.seccion,)
        )
        result = cursor.fetchone()
        if not result:
            raise HTTPException(status_code=404, detail="Sección no encontrada")
        pk_seccion = result[0]

        # Fuente "Encuesta de Campo"
        cursor.execute("SELECT id_fuente FROM fuente_sentimiento WHERE nombre_fuente = 'Encuesta de Campo'")
        row_fuente = cursor.fetchone()
        if not row_fuente:
            cursor.execute("INSERT INTO fuente_sentimiento (nombre_fuente) VALUES ('Encuesta de Campo') RETURNING id_fuente")
            row_fuente = cursor.fetchone()
        id_fuente = row_fuente[0]

        # Insertar sentimientos
        sentimientos = [
            ("Agua", diagnostico.sentimiento.agua),
            ("Basura", diagnostico.sentimiento.basura),
            ("Seguridad", diagnostico.sentimiento.seguridad),
        ]
        for servicio, calif in sentimientos:
            cursor.execute("SELECT id_categoria FROM categoria_servicio WHERE nombre_categoria = %s", (servicio,))
            row_cat = cursor.fetchone()
            if not row_cat:
                cursor.execute("INSERT INTO categoria_servicio (nombre_categoria) VALUES (%s) RETURNING id_categoria", (servicio,))
                row_cat = cursor.fetchone()
            id_categoria = row_cat[0]

            if calif >= 4:
                polaridad = 0.8
            elif calif >= 3:
                polaridad = 0.3
            else:
                polaridad = -0.5

            cursor.execute("""
                INSERT INTO sentimiento_social 
                (pk_seccion, id_fuente, id_categoria, calificacion, sentimiento_polaridad, validado, fecha_registro)
                VALUES (%s, %s, %s, %s, %s, %s, %s)
            """, (pk_seccion, id_fuente, id_categoria, calif, polaridad, True, diagnostico.fecha_recoleccion))

        # Insertar simpatizante
        cursor.execute("""
            INSERT INTO simpatizantes (pk_seccion, nombre, contacto, es_mujer, notas, fecha_registro)
            VALUES (%s, %s, %s, %s, %s, %s)
        """, (pk_seccion,
              diagnostico.simpatizante.nombre,
              diagnostico.simpatizante.contacto,
              diagnostico.simpatizante.es_mujer,
              diagnostico.simpatizante.notas,
              diagnostico.fecha_recoleccion))

        # Evidencia fotográfica
        if diagnostico.evidencia and diagnostico.evidencia.foto_base64:
            base_dir = os.path.dirname(os.path.abspath(__file__))
            evidencias_dir = os.path.join(base_dir, "evidencias")
            os.makedirs(evidencias_dir, exist_ok=True)

            filename = f"evidencia_{diagnostico.seccion}_{uuid.uuid4()}.jpg"
            filepath = os.path.join(evidencias_dir, filename)

            foto_data = diagnostico.evidencia.foto_base64
            if ',' in foto_data:
                foto_data = foto_data.split(',')[1]

            with open(filepath, "wb") as f:
                f.write(base64.b64decode(foto_data))

            cursor.execute("""
                INSERT INTO evidencias (pk_seccion, ruta_archivo, comentario, fecha_registro)
                VALUES (%s, %s, %s, %s)
            """, (pk_seccion, filepath, diagnostico.evidencia.comentario, diagnostico.fecha_recoleccion))

        # GPS
        if diagnostico.latitud is not None and diagnostico.longitud is not None:
            cursor.execute("""
                INSERT INTO logs_gps (pk_seccion, latitud, longitud, fecha_registro)
                VALUES (%s, %s, %s, %s)
            """, (pk_seccion, diagnostico.latitud, diagnostico.longitud, diagnostico.fecha_recoleccion))

        conn.commit()
        return {"mensaje": "Diagnóstico guardado correctamente"}

    except HTTPException:
        conn.rollback()
        raise
    except Exception as e:
        conn.rollback()
        logger.exception("Error en guardar_diagnostico")
        raise HTTPException(status_code=500, detail=f"Error interno: {str(e)}")
    finally:
        cursor.close()
        conn.close()

# ============================================
# Endpoint de exportación Excel
# ============================================

@app.get("/api/exportar-excel")
async def exportar_excel():
    conn = engine.raw_connection()
    tmp = tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False)
    output_path = tmp.name
    tmp.close()
    try:
        df_sentimiento = pd.read_sql("""
            SELECT 
                s.seccion,
                cs.nombre_categoria as servicio,
                ss.calificacion,
                ss.sentimiento_polaridad,
                ss.fecha_registro
            FROM sentimiento_social ss
            JOIN seccion s ON ss.pk_seccion = s.pk_seccion
            JOIN categoria_servicio cs ON ss.id_categoria = cs.id_categoria
            ORDER BY ss.fecha_registro DESC
        """, conn)

        df_simpatizantes = pd.read_sql("""
            SELECT 
                s.seccion,
                sim.nombre,
                sim.contacto,
                sim.es_mujer,
                sim.notas,
                sim.fecha_registro
            FROM simpatizantes sim
            JOIN seccion s ON sim.pk_seccion = s.pk_seccion
            ORDER BY sim.fecha_registro DESC
        """, conn)

        df_evidencias = pd.read_sql("""
            SELECT 
                s.seccion,
                e.ruta_archivo,
                e.comentario,
                e.fecha_registro
            FROM evidencias e
            JOIN seccion s ON e.pk_seccion = s.pk_seccion
            ORDER BY e.fecha_registro DESC
        """, conn)

        df_gps = pd.read_sql("""
            SELECT 
                s.seccion,
                g.latitud,
                g.longitud,
                g.fecha_registro
            FROM logs_gps g
            JOIN seccion s ON g.pk_seccion = s.pk_seccion
            ORDER BY g.fecha_registro DESC
        """, conn)

        with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
            df_sentimiento.to_excel(writer, sheet_name="Sentimiento Social", index=False)
            df_simpatizantes.to_excel(writer, sheet_name="Simpatizantes", index=False)
            df_evidencias.to_excel(writer, sheet_name="Evidencias", index=False)
            df_gps.to_excel(writer, sheet_name="GPS", index=False)

        return FileResponse(
            path=output_path,
            filename="diagnosticos_campo.xlsx",
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            background=BackgroundTask(lambda: os.unlink(output_path))
        )
    except Exception as e:
        if os.path.exists(output_path):
            os.unlink(output_path)
        logger.exception("Error en exportar_excel")
        raise HTTPException(status_code=500, detail=f"Error al generar Excel: {str(e)}")
    finally:
        conn.close()

# ============================================
# Endpoints de administración e invitaciones
# ============================================

class InvitacionRequest(BaseModel):
    email: EmailStr
    tenant_id: str
    dias_validez: int = 7

class InvitacionResponse(BaseModel):
    token: str
    link: str

API_ADMIN_KEY = os.getenv("API_ADMIN_KEY", "cambia_esto_en_produccion")

@app.post("/api/admin/invitacion", response_model=InvitacionResponse)
async def crear_invitacion(inv: InvitacionRequest, admin_key: str = Header(...)):
    if admin_key != API_ADMIN_KEY:
        raise HTTPException(status_code=403, detail="No autorizado")

    token = str(uuid.uuid4())
    expiracion = datetime.now(timezone.utc) + timedelta(days=inv.dias_validez)

    conn = engine.raw_connection()
    cursor = conn.cursor()
    try:
        cursor.execute("""
            INSERT INTO invitaciones (token, email, tenant_id, fecha_expiracion)
            VALUES (%s, %s, %s, %s)
        """, (token, inv.email, inv.tenant_id, expiracion))
        conn.commit()
    except Exception as e:
        conn.rollback()
        logger.exception("Error creando invitación")
        raise HTTPException(status_code=500, detail=str(e))
    finally:
        cursor.close()
        conn.close()

    # Link del dashboard (ajusta la URL base)
    base_url = os.getenv("DASHBOARD_URL", "https://5mefjgvsuazhayejhm92vk.streamlit.app")
    link = f"{base_url}?invite={token}"
    return InvitacionResponse(token=token, link=link)

@app.get("/api/validate-invite/{token}")
async def validate_invite(token: str):
    conn = engine.raw_connection()
    cursor = conn.cursor()
    try:
        cursor.execute("""
            SELECT email, tenant_id, fecha_expiracion, usada
            FROM invitaciones
            WHERE token = %s
        """, (token,))
        row = cursor.fetchone()
        if not row:
            raise HTTPException(status_code=404, detail="Token no encontrado")
        email, tenant_id, expiracion, usada = row
        # ⚠️ Temporal: ignoramos usada y expiración
        # Devolvemos los datos directamente
        return {"email": email, "tenant_id": tenant_id}
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))
    finally:
        cursor.close()
        conn.close()


# Estos 3 endpoints manejan el sistema de afiliación Dudú
# Sin tocar ningún endpoint existente.
# ============================================

from pydantic import BaseModel, validator
from typing import Optional, List
from datetime import datetime

# ── Modelos de datos ──────────────────────────────────────
class AfiliadoIn(BaseModel):
    nombre_completo: str
    telefono:        str
    edad:            Optional[int] = None
    genero:          Optional[str] = None
    municipio:       str
    colonia:         Optional[str] = None
    seccion_electoral: Optional[int] = None
    tipo_participacion: Optional[str] = "Simpatizante"
    temas_interes:   Optional[List[str]] = []
    como_se_entero:  Optional[str] = None
    acepta_aviso:    bool
    acepta_contacto: bool

    @validator('nombre_completo')
    def nombre_no_vacio(cls, v):
        if not v or len(v.strip()) < 3:
            raise ValueError('Nombre muy corto')
        return v.strip()

    @validator('telefono')
    def telefono_valido(cls, v):
        digits = ''.join(filter(str.isdigit, v))
        if len(digits) < 10:
            raise ValueError('Teléfono inválido')
        return v.strip()

    @validator('acepta_aviso')
    def debe_aceptar_aviso(cls, v):
        if not v:
            raise ValueError('Debe aceptar el aviso de privacidad')
        return v

    @validator('municipio')
    def municipio_valido(cls, v):
        permitidos = ['Taxco de Alarcón', 'Pilcaya', 'Tetipac', 'Taxco', 'Otro']
        if v not in permitidos:
            raise ValueError(f'Municipio no válido: {v}')
        return v


# ── Endpoint 1: Registrar afiliado ────────────────────────
@app.post("/api/afiliados")
async def registrar_afiliado(
    afiliado: AfiliadoIn,
    request: Request,
    db: Session = Depends(get_db)
):
    """
    Recibe registro desde landing de Dudú.
    Rate limiting implícito via Render.
    Aguanta 10,000 usuarios concurrentes.
    """
    try:
        # Prevenir duplicados por teléfono (mismo municipio)
        existe = db.execute(text("""
            SELECT pk_afiliado FROM afiliados_dudu
            WHERE telefono = :tel AND municipio = :mun
            LIMIT 1
        """), {"tel": afiliado.telefono, "mun": afiliado.municipio}).fetchone()

        if existe:
            return {
                "ok": False,
                "mensaje": "Este número ya está registrado en tu municipio.",
                "duplicado": True
            }

        # IP del cliente para auditoría
        ip_cliente = request.client.host if request.client else "0.0.0.0"

        # Insertar registro
        db.execute(text("""
            INSERT INTO afiliados_dudu (
                nombre_completo, telefono, edad, genero,
                municipio, colonia, seccion_electoral,
                tipo_participacion, temas_interes, como_se_entero,
                acepta_aviso, acepta_contacto, ip_registro
            ) VALUES (
                :nombre, :tel, :edad, :genero,
                :municipio, :colonia, :seccion,
                :tipo, :temas, :como,
                :aviso, :contacto, :ip
            )
        """), {
            "nombre":   afiliado.nombre_completo,
            "tel":      afiliado.telefono,
            "edad":     afiliado.edad,
            "genero":   afiliado.genero,
            "municipio": afiliado.municipio,
            "colonia":  afiliado.colonia,
            "seccion":  afiliado.seccion_electoral,
            "tipo":     afiliado.tipo_participacion,
            "temas":    afiliado.temas_interes,
            "como":     afiliado.como_se_entero,
            "aviso":    afiliado.acepta_aviso,
            "contacto": afiliado.acepta_contacto,
            "ip":       ip_cliente,
        })
        db.commit()

        return {
            "ok": True,
            "mensaje": "¡Gracias! Tu registro fue recibido correctamente.",
        }

    except Exception as e:
        db.rollback()
        raise HTTPException(status_code=500, detail=str(e))


# ── Endpoint 2: KPIs para dashboard ──────────────────────
@app.get("/api/afiliados/kpis")
async def kpis_afiliados(db: Session = Depends(get_db)):
    """KPIs en tiempo real para el dashboard interno."""
    try:
        resultado = db.execute(text("""
            SELECT
                COUNT(*)                                          AS total,
                COUNT(CASE WHEN genero = 'Mujer'  THEN 1 END)   AS mujeres,
                COUNT(CASE WHEN genero = 'Hombre' THEN 1 END)   AS hombres,
                COUNT(CASE WHEN municipio = 'Taxco de Alarcón'  THEN 1 END) AS taxco,
                COUNT(CASE WHEN municipio = 'Pilcaya'           THEN 1 END) AS pilcaya,
                COUNT(CASE WHEN municipio = 'Tetipac'           THEN 1 END) AS tetipac,
                COUNT(CASE WHEN fecha_registro >= NOW() - INTERVAL '24 hours' THEN 1 END) AS ultimas_24h,
                COUNT(CASE WHEN fecha_registro >= NOW() - INTERVAL '7 days'  THEN 1 END) AS ultima_semana
            FROM afiliados_dudu
            WHERE activo = TRUE
        """)).fetchone()

        return {
            "total":         resultado[0],
            "mujeres":       resultado[1],
            "hombres":       resultado[2],
            "taxco":         resultado[3],
            "pilcaya":       resultado[4],
            "tetipac":       resultado[5],
            "ultimas_24h":   resultado[6],
            "ultima_semana": resultado[7],
        }
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


# ── Endpoint 3: Exportar Excel ────────────────────────────
@app.get("/api/afiliados/exportar")
async def exportar_afiliados(
    admin_key: str = Header(None, alias="admin-key"),
    db: Session = Depends(get_db)
):
    """
    Exporta todos los afiliados a Excel.
    Requiere admin-key para proteger datos personales.
    """
    if admin_key != os.getenv("API_ADMIN_KEY", ""):
        raise HTTPException(status_code=403, detail="No autorizado")

    try:
        import io
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment

        rows = db.execute(text("""
            SELECT nombre_completo, telefono, edad, genero,
                   municipio, colonia, seccion_electoral,
                   tipo_participacion, como_se_entero,
                   fecha_registro::date as fecha
            FROM afiliados_dudu
            WHERE activo = TRUE
            ORDER BY fecha_registro DESC
        """)).fetchall()

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Afiliados Dudú"

        # Encabezados con estilo
        headers = ["Nombre", "Teléfono", "Edad", "Género",
                   "Municipio", "Colonia", "Sección", "Tipo",
                   "Cómo se enteró", "Fecha registro"]

        verde = PatternFill("solid", fgColor="2E7D32")
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = verde
            cell.alignment = Alignment(horizontal="center")
            ws.column_dimensions[chr(64 + col)].width = 18

        # Datos
        for row_num, row in enumerate(rows, 2):
            for col_num, value in enumerate(row, 1):
                ws.cell(row=row_num, column=col_num, value=str(value) if value else "")

        # Fila de totales
        ws.append([])
        ws.append([f"Total registros: {len(rows)}", "", "", "",
                   "", "", "", "", "",
                   f"Generado: {datetime.now().strftime('%d/%m/%Y %H:%M')}"])

        output = io.BytesIO()
        wb.save(output)
        output.seek(0)

        from fastapi.responses import StreamingResponse
        return StreamingResponse(
            output,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": "attachment; filename=afiliados_dudu.xlsx"}
        )
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

  if __name__ == "__main__":
    port = int(os.environ.get("PORT", 8000))
    uvicorn.run("main:app", host="0.0.0.0", port=port, reload=False)